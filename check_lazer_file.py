# -*- coding: utf8 -*-

from openpyxl import load_workbook
wb = load_workbook(r'C:\Users\АПК\Dropbox\tehotdel\Реестр заказов резки цех\реестр заказов резки цех.xlsx')
ws = wb.active
print('Невыполненные заявки:')
print()
flag = False
i = 700
while ws['A' + str(i)].value != None:
    if ws['C' + str(i)].value == 'Агарков' and ws['J' + str(i)].value == None:
        print(ws['A' + str(i)].value, str(ws['I' + str(i)].value)[:-9], ws['C' + str(i)].value, ws['D' + str(i)].value, sep='   ')
        flag = True
    i += 1
if flag == False:
    print('\t', 'Отсутствуют')
print()
input("Нажмите Enter, чтобы закрыть")