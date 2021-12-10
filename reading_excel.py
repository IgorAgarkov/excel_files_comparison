# -*- coding: utf8 -*-

from openpyxl import load_workbook
wb = load_workbook(r'C:\Users\АПК\Dropbox\tehotdel\Реестр заказов резки цех\реестр заказов резки цех.xlsx')
ws = wb.active
c = ws['J1001'].value
if c != None:
    print('Отсекатель запущен в работу, можно нести сборочный в цех')
    input('Нажмите Enter для выхода')